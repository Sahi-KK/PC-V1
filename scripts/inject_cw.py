#!/usr/bin/env python3
import os
import re
import requests
import openpyxl
from datetime import datetime
from dotenv import load_dotenv

load_dotenv('.env.local')

SUPABASE_URL = os.getenv('NEXT_PUBLIC_SUPABASE_URL')
SERVICE_KEY  = os.getenv('SUPABASE_SERVICE_ROLE_KEY')

HEADERS = {
    'apikey': SERVICE_KEY,
    'Authorization': f'Bearer {SERVICE_KEY}',
    'Content-Type': 'application/json',
    'Prefer': 'return=minimal',
}

def supabase_upsert(table: str, data: list, conflict_col: str = None):
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    prefer = 'return=minimal'
    if conflict_col:
        prefer += f',resolution=merge-duplicates'
    headers = {**HEADERS, 'Prefer': prefer}
    params = {}
    if conflict_col:
        params['on_conflict'] = conflict_col
    resp = requests.post(url, headers=headers, json=data, params=params)
    if resp.status_code not in (200, 201):
        print(f"  ERROR {table}: {resp.status_code} {resp.text[:300]}")
        return False
    return True

def parse_class_code(code: str):
    code = str(code).strip()
    m = re.match(r'^(.+?)\s+(\d+)\s*\(', code)
    if m:
        abbr = m.group(1).strip()
        session = int(m.group(2))
        return abbr, session
    return code, 1

def main():
    print("Loading new file...")
    new_wb = openpyxl.load_workbook('../Academic Calendar and Courses Term IV - PC V1.xlsx', data_only=True)

    # 1. Course List
    courses_data = []
    for row in new_wb['Course List'].iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        abbr = str(row[0]).strip()
        if abbr != 'CW-':
            continue
            
        faculty = str(row[1]).strip() if row[1] else ''
        faculty_abbr = str(row[2]).strip() if row[2] else ''
        credit = float(row[3]) if row[3] else 0
        full_name = 'Communication Workshop'
        
        courses_data.append({
            'abbr': abbr,
            'full_name': full_name,
            'faculty': faculty,
            'faculty_abbr': faculty_abbr,
            'credit': credit,
        })
        break

    if courses_data:
        print(f"Upserting Course: {courses_data[0]['abbr']}")
        supabase_upsert('courses', courses_data, 'abbr')

    # 2. Students and Courses (Enrollments)
    print("Fetching existing student IDs from DB...")
    resp = requests.get(
        f"{SUPABASE_URL}/rest/v1/students",
        headers=HEADERS,
        params={'select': 'id,roll_no'}
    )
    student_id_map = {s['roll_no']: s['id'] for s in resp.json()}

    enrollments = []
    seen = set()
    for row in new_wb['Master 1'].iter_rows(min_row=2, values_only=True):
        if row[1] is None or row[5] is None:
            continue
        roll = str(row[1]).strip()
        course = str(row[5]).strip()
        if course == 'CW-':
            sid = student_id_map.get(roll)
            if not sid:
                continue
            section = str(row[6]).strip() if row[6] else 'A'
            key = (sid, course)
            if key not in seen:
                seen.add(key)
                enrollments.append({
                    'student_id': sid,
                    'course_abbr': course,
                    'section': section,
                })

    if enrollments:
        print(f"Upserting {len(enrollments)} CW- enrollments...")
        for i in range(0, len(enrollments), 500):
            batch = enrollments[i:i+500]
            supabase_upsert('student_courses', batch, 'student_id,course_abbr')

    # 3. Calendar Entries
    SECTION_LR = {'A': 'LR 02', 'B': 'LR 07', 'C': 'LR 06', 'D': 'LR 06'}
    TIME_SLOTS = [
        '08:45-10:00', '10:20-11:35', '11:55-1:10', 'LUNCH',
        '14:30-15:45', '16:05-17:20', '17:40-18:55',
        '19:15-20:30', '20:50-22:05', '22:25-23:40'
    ]
    EXAM_START = '2026-08-23'
    INDEPENDENCE_DAY = '2026-08-15'

    cal_entries = []
    current_date = None

    for row in new_wb['PGPO Term IV'].iter_rows(min_row=4, values_only=True):
        col0 = row[0]
        col1 = row[1]
        slots = list(row[2:12])

        if col0 is not None and isinstance(col0, datetime):
            current_date = col0.date().isoformat()
            
        if not current_date:
            continue

        is_holiday = current_date == INDEPENDENCE_DAY
        is_exam = current_date >= EXAM_START

        if col1 and isinstance(col1, str) and col1 in ['A', 'B', 'C', 'D']:
            section = col1
            lr = SECTION_LR.get(section, '')
            for j, slot_time in enumerate(TIME_SLOTS):
                if slot_time == 'LUNCH':
                    continue
                if j < len(slots) and slots[j]:
                    raw = str(slots[j]).strip()
                    if raw.upper() in ('SUNDAY', 'HOLIDAY', ''):
                        continue
                    
                    course_abbr, session_num = parse_class_code(raw)
                    if course_abbr == 'CW-':
                        cal_entries.append({
                            'date': current_date,
                            'day_of_week': None,
                            'section': section,
                            'lr': lr,
                            'time_slot': slot_time,
                            'class_code': raw,
                            'course_abbr': course_abbr,
                            'session_number': session_num,
                            'note': None,
                            'is_holiday': is_holiday,
                            'is_exam_period': is_exam,
                        })

    print(f"Found {len(cal_entries)} CW- calendar entries.")
    
    # We must fetch existing CW- entries so we don't duplicate them
    print("Checking for existing CW- calendar entries in DB...")
    resp = requests.get(
        f"{SUPABASE_URL}/rest/v1/calendar_entries",
        headers=HEADERS,
        params={'select': 'date,time_slot,class_code,section', 'course_abbr': 'eq.CW-'}
    )
    existing_cal = set()
    for e in resp.json():
        existing_cal.add(f"{e['date']}|{e['section']}|{e['time_slot']}|{e['class_code']}")

    new_cal_entries = []
    for e in cal_entries:
        key = f"{e['date']}|{e['section']}|{e['time_slot']}|{e['class_code']}"
        if key not in existing_cal:
            new_cal_entries.append(e)

    if new_cal_entries:
        print(f"Upserting {len(new_cal_entries)} NEW CW- calendar entries to Supabase...")
        success = supabase_upsert('calendar_entries', new_cal_entries)
        if success:
            print("Successfully injected calendar entries!")
    else:
        print("No new CW- calendar entries to inject.")

if __name__ == '__main__':
    main()
