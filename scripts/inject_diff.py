#!/usr/bin/env python3
import os
import re
import requests
import openpyxl
from datetime import datetime
from dotenv import load_dotenv

load_dotenv('.env.local')

SUPABASE_URL = os.getenv('NEXT_PUBLIC_SUPABASE_URL')
SERVICE_KEY  = os.getenv('SUPABASE_SERVICE_ROLE_KEY')

HEADERS = {
    'apikey': SERVICE_KEY,
    'Authorization': f'Bearer {SERVICE_KEY}',
    'Content-Type': 'application/json',
    'Prefer': 'return=minimal',
}

def supabase_upsert(table: str, data: list):
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    resp = requests.post(url, headers=HEADERS, json=data)
    if resp.status_code not in (200, 201):
        print(f"  ERROR {table}: {resp.status_code} {resp.text[:300]}")
        return False
    return True

def parse_class_code(code: str):
    code = str(code).strip()
    m = re.match(r'^(.+?)\s+(\d+)\s*\(', code)
    if m:
        abbr = m.group(1).strip()
        session = int(m.group(2))
        return abbr, session
    return code, 1

def main():
    print("Loading old file...")
    old_wb = openpyxl.load_workbook('../Academic Calendar and Courses.xlsx', data_only=True)
    old_slots = set()
    for row in old_wb['Term IV Calendar'].iter_rows(min_row=4, values_only=True):
        date = row[0]
        section = row[1]
        for j, val in enumerate(row[2:12]):
            if val and str(val).strip():
                old_slots.add(f"{date}|{section}|{j}|{str(val).strip()}")

    print("Loading new file...")
    new_wb = openpyxl.load_workbook('../Academic Calendar and Courses Term IV - PC V1.xlsx', data_only=True)

    SECTION_LR = {'A': 'LR 02', 'B': 'LR 07', 'C': 'LR 06', 'D': 'LR 06'}
    TIME_SLOTS = [
        '08:45-10:00', '10:20-11:35', '11:55-1:10', 'LUNCH',
        '14:30-15:45', '16:05-17:20', '17:40-18:55',
        '19:15-20:30', '20:50-22:05', '22:25-23:40'
    ]
    EXAM_START = '2026-08-23'
    INDEPENDENCE_DAY = '2026-08-15'

    cal_entries = []
    current_date = None

    for row in new_wb['PGPO Term IV'].iter_rows(min_row=4, values_only=True):
        col0 = row[0]
        col1 = row[1]
        slots = list(row[2:12])

        if col0 is not None and isinstance(col0, datetime):
            current_date = col0.date().isoformat()
            
        if not current_date:
            continue

        is_holiday = current_date == INDEPENDENCE_DAY
        is_exam = current_date >= EXAM_START

        if col1 and isinstance(col1, str) and col1 in ['A', 'B', 'C', 'D']:
            section = col1
            lr = SECTION_LR.get(section, '')
            for j, slot_time in enumerate(TIME_SLOTS):
                if slot_time == 'LUNCH':
                    continue
                if j < len(slots) and slots[j]:
                    raw = str(slots[j]).strip()
                    if raw.upper() in ('SUNDAY', 'HOLIDAY', ''):
                        continue
                    
                    key = f"{row[0]}|{section}|{j}|{raw}"
                    if key not in old_slots:
                        course_abbr, session_num = parse_class_code(raw)
                        cal_entries.append({
                            'date': current_date,
                            'day_of_week': None,
                            'section': section,
                            'lr': lr,
                            'time_slot': slot_time,
                            'class_code': raw,
                            'course_abbr': course_abbr,
                            'session_number': session_num,
                            'note': None,
                            'is_holiday': is_holiday,
                            'is_exam_period': is_exam,
                        })

    print(f"Found {len(cal_entries)} NEW calendar entries.")
    for e in cal_entries:
        print(f"  {e['date']} {e['time_slot']} {e['class_code']}")
        
    if cal_entries:
        print("Upserting new calendar entries to Supabase...")
        success = supabase_upsert('calendar_entries', cal_entries)
        if success:
            print("Successfully injected!")

if __name__ == '__main__':
    main()
