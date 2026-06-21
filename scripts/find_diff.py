import openpyxl

def main():
    print("Loading old file...")
    old_wb = openpyxl.load_workbook('../Academic Calendar and Courses.xlsx', data_only=True)
    print("Loading new file...")
    new_wb = openpyxl.load_workbook('../Academic Calendar and Courses Term IV - PC V1.xlsx', data_only=True)

    # 1. Compare Course List
    old_courses = set()
    for row in old_wb['Course List'].iter_rows(min_row=2, values_only=True):
        if row[0]:
            old_courses.add(str(row[0]).strip())

    new_courses = set()
    for row in new_wb['Course List'].iter_rows(min_row=2, values_only=True):
        if row[0]:
            new_courses.add(str(row[0]).strip())

    added_courses = new_courses - old_courses
    print(f"Added Courses: {added_courses}")

    # 2. Compare Calendar Slots
    old_slots = set()
    for i, row in enumerate(old_wb['Term IV Calendar'].iter_rows(min_row=4, values_only=True)):
        date = row[0]
        section = row[1]
        for j, val in enumerate(row[2:12]):
            if val and str(val).strip():
                old_slots.add(f"{date}|{section}|{j}|{str(val).strip()}")

    new_slots = []
    for i, row in enumerate(new_wb['PGPO Term IV'].iter_rows(min_row=4, values_only=True)):
        date = row[0]
        section = row[1]
        for j, val in enumerate(row[2:12]):
            if val and str(val).strip():
                key = f"{date}|{section}|{j}|{str(val).strip()}"
                if key not in old_slots:
                    new_slots.append(key)

    print(f"Added Calendar Slots: {len(new_slots)}")
    for s in new_slots:
        print("  " + s)

    # 3. Compare Students enrollments
    old_enrollments = set()
    for row in old_wb['Students and Courses'].iter_rows(min_row=2, values_only=True):
        if row[1] and row[4]:
            old_enrollments.add(f"{str(row[1]).strip()}|{str(row[4]).strip()}|{str(row[5]).strip() if row[5] else 'A'}")

    new_enrollments = set()
    for row in new_wb['Master 1'].iter_rows(min_row=2, values_only=True):
        if row[1] and row[5]:
            new_enrollments.add(f"{str(row[1]).strip()}|{str(row[5]).strip()}|{str(row[6]).strip() if row[6] else 'A'}")

    added_enrollments = new_enrollments - old_enrollments
    print(f"Added Enrollments: {len(added_enrollments)}")
    for e in added_enrollments:
        print("  " + e)

if __name__ == '__main__':
    main()
