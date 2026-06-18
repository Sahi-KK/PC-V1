#!/usr/bin/env python3
"""
Seed Supabase database with data from the Academic Calendar Excel file.
Run: python3 scripts/seed_database.py

Requirements:
  pip install openpyxl requests python-dotenv
"""

import os
import json
import re
import requests
from datetime import datetime
from dotenv import load_dotenv

load_dotenv('../.env.local')

SUPABASE_URL = os.getenv('NEXT_PUBLIC_SUPABASE_URL')
SERVICE_KEY  = os.getenv('SUPABASE_SERVICE_ROLE_KEY')

if not SUPABASE_URL or not SERVICE_KEY:
    raise ValueError("Missing SUPABASE_URL or SERVICE_KEY in .env.local")

HEADERS = {
    'apikey': SERVICE_KEY,
    'Authorization': f'Bearer {SERVICE_KEY}',
    'Content-Type': 'application/json',
    'Prefer': 'return=minimal',
}

def supabase_upsert(table: str, data: list, conflict_col: str = None):
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    prefer = 'return=minimal'
    if conflict_col:
        prefer += f',resolution=merge-duplicates'
    headers = {**HEADERS, 'Prefer': prefer}
    params = {}
    if conflict_col:
        params['on_conflict'] = conflict_col
    resp = requests.post(url, headers=headers, json=data, params=params)
    if resp.status_code not in (200, 201):
        print(f"  ERROR {table}: {resp.status_code} {resp.text[:300]}")
        return False
    return True


def main():
    import openpyxl

    print("Loading Excel file...")
    wb = openpyxl.load_workbook(
        '../Academic Calendar and Courses.xlsx',
        data_only=True
    )

    # ────────────────────────────────
    # 1. Parse Course List
    # ────────────────────────────────
    ws_courses = wb['Course List']
    ws_cal = wb['Term IV Calendar']

    # Build full_name map from calendar sheet (cols 14-15 = full name, abbr)
    full_name_map = {}
    for row in ws_cal.iter_rows(min_row=3, values_only=True):
        if row[13] and row[14]:
            abbr = str(row[14]).strip()
            full_name = str(row[13]).strip()
            # Clean up – remove trailing (ABBR)
            full_name_map[abbr] = full_name

    # Add L&D manually (appears with & in sheet)
    full_name_map['L&D'] = 'Learning and Development (L&D)'

    courses_data = []
    for row in ws_courses.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        abbr = str(row[0]).strip()
        if abbr == 'CW-':
            continue  # skip CW for now
        faculty = str(row[1]).strip() if row[1] else ''
        faculty_abbr = str(row[2]).strip() if row[2] else ''
        credit = float(row[3]) if row[3] else 0
        full_name = full_name_map.get(abbr, abbr)

        courses_data.append({
            'abbr': abbr,
            'full_name': full_name,
            'faculty': faculty,
            'faculty_abbr': faculty_abbr,
            'credit': credit,
        })

    print(f"Upserting {len(courses_data)} courses...")
    supabase_upsert('courses', courses_data, 'abbr')

    # ────────────────────────────────
    # 2. Parse Students
    # ────────────────────────────────
    ws_students = wb['Students and Courses']
    students_map = {}  # roll_no → {name, courses: [(abbr, section)]}

    for row in ws_students.iter_rows(min_row=2, values_only=True):
        if row[2] is None:
            continue
        roll = str(row[1]).strip() if row[1] else None
        name = str(row[2]).strip() if row[2] else None
        course = str(row[4]).strip() if row[4] else None
        section = str(row[5]).strip() if row[5] else None

        if not roll or not name:
            continue

        if roll not in students_map:
            students_map[roll] = {'name': name, 'roll_no': roll, 'courses': []}

        if course and course != 'CW-':
            students_map[roll]['courses'].append((course, section))

    students_list = [{'roll_no': v['roll_no'], 'name': v['name']} for v in students_map.values()]
    print(f"Upserting {len(students_list)} students...")
    supabase_upsert('students', students_list, 'roll_no')

    # Fetch student IDs
    resp = requests.get(
        f"{SUPABASE_URL}/rest/v1/students",
        headers=HEADERS,
        params={'select': 'id,roll_no'}
    )
    student_id_map = {s['roll_no']: s['id'] for s in resp.json()}

    # ────────────────────────────────
    # 3. Student Course enrollments
    # ────────────────────────────────
    enrollments = []
    for roll, data in students_map.items():
        sid = student_id_map.get(roll)
        if not sid:
            continue
        seen = set()
        for (course_abbr, section) in data['courses']:
            key = (sid, course_abbr)
            if key in seen:
                continue
            seen.add(key)
            enrollments.append({
                'student_id': sid,
                'course_abbr': course_abbr,
                'section': section or 'A',
            })

    print(f"Upserting {len(enrollments)} enrollments...")
    # Insert in batches of 500
    for i in range(0, len(enrollments), 500):
        batch = enrollments[i:i+500]
        supabase_upsert('student_courses', batch, 'student_id,course_abbr')

    # ────────────────────────────────
    # 4. Parse Calendar
    # ────────────────────────────────
    SECTION_LR = {'A': 'LR 02', 'B': 'LR 07', 'C': 'LR 06', 'D': 'LR 06'}
    TIME_SLOTS = [
        '08:45-10:00', '10:20-11:35', '11:55-1:10', 'LUNCH',
        '14:30-15:45', '16:05-17:20', '17:40-18:55',
        '19:15-20:30', '20:50-22:05', '22:25-23:40'
    ]

    EXAM_START = '2026-08-23'
    INDEPENDENCE_DAY = '2026-08-15'

    def parse_class_code(code: str):
        """Parse 'CV 1 (SA)' → course_abbr='CV', session_number=1"""
        code = str(code).strip()
        # Match patterns like 'MBPET 20 (RN)' or 'L&D 5 (SY)' or 'Ind 4.0 3 (AK1)'
        m = re.match(r'^(.+?)\s+(\d+)\s*\(', code)
        if m:
            abbr = m.group(1).strip()
            session = int(m.group(2))
            return abbr, session
        return code, 1

    cal_entries = []
    current_date = None
    DAY_NAMES = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']

    for i, row in enumerate(ws_cal.iter_rows(min_row=4, values_only=True)):
        col0 = row[0]
        col1 = row[1]
        slots = list(row[2:12])

        # Update current date
        if col0 is not None and isinstance(col0, datetime):
            current_date = col0.date().isoformat()

        if not current_date:
            continue

        is_holiday = current_date == INDEPENDENCE_DAY
        is_exam = current_date >= EXAM_START

        if col1 and isinstance(col1, str) and col1 in ['A', 'B', 'C', 'D']:
            section = col1
            lr = SECTION_LR.get(section, '')
            for j, slot_time in enumerate(TIME_SLOTS):
                if slot_time == 'LUNCH':
                    continue
                if j < len(slots) and slots[j]:
                    raw = str(slots[j]).strip()
                    if raw.upper() in ('SUNDAY', 'HOLIDAY', ''):
                        continue
                    course_abbr, session_num = parse_class_code(raw)

                    cal_entries.append({
                        'date': current_date,
                        'day_of_week': None,  # will be derived client-side
                        'section': section,
                        'lr': lr,
                        'time_slot': slot_time,
                        'class_code': raw,
                        'course_abbr': course_abbr,
                        'session_number': session_num,
                        'note': None,
                        'is_holiday': is_holiday,
                        'is_exam_period': is_exam,
                    })

    print(f"Upserting {len(cal_entries)} calendar entries...")
    for i in range(0, len(cal_entries), 500):
        batch = cal_entries[i:i+500]
        supabase_upsert('calendar_entries', batch)

    # ────────────────────────────────
    # 5. Allowed users
    # ────────────────────────────────
    allowed_users = [
        {
            'email': 'ipm04krishnakants@iimrohtak.ac.in',
            'roll_no': 'IPM04134',
            'name': 'Krishnakant Singh',
        }
    ]
    print(f"Upserting {len(allowed_users)} allowed users...")
    supabase_upsert('allowed_users', allowed_users, 'email')

    print("\n✅ Database seeded successfully!")
    print(f"  Courses: {len(courses_data)}")
    print(f"  Students: {len(students_list)}")
    print(f"  Enrollments: {len(enrollments)}")
    print(f"  Calendar entries: {len(cal_entries)}")
    print(f"  Allowed users: {len(allowed_users)}")


if __name__ == '__main__':
    main()
